import openpyxl

wk = openpyxl.load_workbook("C://Users//geefu//OneDrive//Documents//TestData.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row),int(cell))
    return cell.value


# print(fetch_number_of_rows("GeeSlimmy"))
# print(fetch_cell_data("GeeSlimmy",1,1))


# sh = wk["GeeSlimmy"]
# print(sh.max_row)
# cell = sh.cell(1,1)
# print(cell.value)
