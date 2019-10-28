import openpyxl

# Load Workbook

wk = openpyxl.load_workbook("C:\\Users\\geefu\\OneDrive\\Documents\\TestData.xlsx")

print(wk.sheetnames)
print("Active sheet is " + wk.active.title)

# Create object of any sheet on which you want to work

sh=wk['GeeSlimmy']
print(sh.title)

# Read Data from cells

print(sh['A3'].value)
print(sh['B4'].value)

c1= sh.cell(column=3,row=2)
print(c1.value)
print(c1.column)

# Find rows having data

rows = sh.max_row
columns = sh.max_column
print("Total Rows are - " + str(rows))
print("Total Columns are - " + str(columns))

# Syntax for fetching all excel data
for i in range(1,rows+1):
    for j in range(1,columns+1):
        c=sh.cell(i,j)
        print(c.value)

# Syntax 2 for fetching all excel data
for r in sh['A1':'C3']:
    for c in r:
        print(c.value)
